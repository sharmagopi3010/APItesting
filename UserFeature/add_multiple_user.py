import openpyxl
import requests,json

baseURL = "https://thetestingworldapi.com/"
def test_createUser():
    # adding student details
    student_detail_url = baseURL + "api/studentsDetails/"
    f = open('studentdetails.json')
    wk = openpyxl.load_workbook('student_data.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row
    print(rows)

    json_payload = json.loads(f.read())

    for i in range(2, rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)
        json_payload['first_name'] = cell_first_name.value
        json_payload['middle_name'] = cell_middle_name.value
        json_payload['last_name'] = cell_last_name.value
        json_payload['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(student_detail_url, json_payload)
        print(response.status_code)
        response_json = json.loads(response.text)
        print(response_json)